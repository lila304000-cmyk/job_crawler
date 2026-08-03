"""数据标准化服务。

功能：
1. 读取企业、岗位、技能三套导入模板；
2. 复用 reference_spider 的清洗逻辑（去 HTML/空白、广告文案、正则提取）；
3. 按岗位链接去重、统一日期薪资格式、技能匹配标准库；
4. 输出严格对齐两套导入模板的结构化数据。
"""

from __future__ import annotations

import asyncio
import hashlib
import html
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models import Channel, CrawlLog, JobRecord, Task


# ==================== 常量定义 ====================

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
TEMPLATES_DIR = PROJECT_ROOT / "data" / "templates"

COMPANY_TEMPLATE_FILE = TEMPLATES_DIR / "platform-company-import-template.xlsx"
SKILLS_TEMPLATE_FILE = TEMPLATES_DIR / "skills_data.xlsx"
TASK_TEMPLATE_FILE = TEMPLATES_DIR / "task-import-template.xlsx"

MAX_DESC_LENGTH = 5000
MAX_TITLE_LENGTH = 500
MAX_COMPANY_LENGTH = 255

# 货币代码映射
CURRENCY_MAP: Dict[str, str] = {
    "cny": "CNY", "rmb": "CNY", "¥": "CNY", "元": "CNY", "人民币": "CNY",
    "usd": "USD", "$": "USD", "美元": "USD",
    "eur": "EUR", "€": "EUR", "欧元": "EUR",
    "gbp": "GBP", "£": "GBP", "英镑": "GBP",
    "jpy": "JPY", "￥": "JPY", "日元": "JPY",
    "hkd": "HKD", "港币": "HKD", "港元": "HKD",
    "sgd": "SGD", "新币": "SGD", "新加坡元": "SGD",
    "aud": "AUD", "澳元": "AUD",
    "cad": "CAD", "加元": "CAD",
}

# 国家代码/名称映射（常用）
COUNTRY_MAP: Dict[str, str] = {
    "cn": "CN", "china": "CN", "中国": "CN", "mainland": "CN",
    "sg": "SG", "singapore": "SG", "新加坡": "SG",
    "us": "US", "usa": "US", "united states": "US", "美国": "US",
    "uk": "GB", "united kingdom": "GB", "britain": "GB", "英国": "GB",
    "jp": "JP", "japan": "JP", "日本": "JP",
    "kr": "KR", "korea": "KR", "south korea": "KR", "韩国": "KR",
    "de": "DE", "germany": "DE", "德国": "DE",
    "fr": "FR", "france": "FR", "法国": "FR",
    "au": "AU", "australia": "AU", "澳大利亚": "AU", "澳洲": "AU",
    "ca": "CA", "canada": "CA", "加拿大": "CA",
    "my": "MY", "malaysia": "MY", "马来西亚": "MY",
    "th": "TH", "thailand": "TH", "泰国": "TH",
    "vn": "VN", "vietnam": "VN", "越南": "VN",
    "id": "ID", "indonesia": "ID", "印尼": "ID",
    "ph": "PH", "philippines": "PH", "菲律宾": "PH",
    "in": "IN", "india": "IN", "印度": "IN",
    "br": "BR", "brazil": "BR", "巴西": "BR",
    "mx": "MX", "mexico": "MX", "墨西哥": "MX",
    "ru": "RU", "russia": "RU", "俄罗斯": "RU",
    "sa": "SA", "saudi arabia": "SA", "沙特阿拉伯": "SA",
    "ae": "AE", "uae": "AE", "united arab emirates": "AE", "阿联酋": "AE",
    "remote": "REMOTE", "线上": "REMOTE", "远程": "REMOTE",
}

# 语言代码映射（与模板示例对齐）
LANGUAGE_MAP: Dict[str, str] = {
    "中文（简体）": "zh-Hans", "中文(简体)": "zh-Hans", "chinese (simplified)": "zh-Hans",
    "简体中文": "zh-Hans", "简体": "zh-Hans", "中文": "zh-Hans", "chinese": "zh-Hans",
    "中文（繁体）": "zh-Hant", "中文(繁体)": "zh-Hant", "繁体中文": "zh-Hant",
    "英语": "en", "英文": "en", "english": "en",
    "日语": "ja", "日文": "ja", "japanese": "ja",
    "韩语": "ko", "韩文": "ko", "korean": "ko",
    "法语": "fr", "法文": "fr", "french": "fr",
    "德语": "de", "德文": "de", "german": "de",
    "西班牙语": "es", "西语": "es", "spanish": "es",
    "葡萄牙语": "pt", "portuguese": "pt",
    "俄语": "ru", "russian": "ru",
    "阿拉伯语": "ar", "arabic": "ar",
    "印尼语": "id", "indonesian": "id",
    "泰语": "th", "thai": "th",
    "越南语": "vi", "vietnamese": "vi",
    "马来语": "ms", "malay": "ms",
}

# 行业关键词 -> 模板行业代码
INDUSTRY_KEYWORDS: Dict[str, List[str]] = {
    "technology": ["tech", "software", "互联网", "科技", "it", "saas", "ai", "人工智能", "data", "data", "cloud", "cloud", "fintech", "blockchain", "crypto", "web3", "电商", "e-commerce", "电子商务", "游戏", "gaming"],
    "finance": ["finance", "banking", "investment", "证券", "基金", "保险", "insurance", "fintech", "financial", "投行", "财富", "支付", "payment"],
    "healthcare": ["health", "medical", "healthcare", "医疗", "医药", "生物科技", "biotech", "hospital", "诊所"],
    "education": ["education", "教育", "培训", "teaching", "school", "university", "在线课程", "e-learning"],
    "manufacturing": ["manufacturing", "制造", "工业", "factory", "production", "automotive", "汽车", "hardware", "硬件", "电子"],
    "retail": ["retail", "零售", "消费品", "consumer", "fmcg", "快消", "贸易", "trading"],
    "logistics": ["logistics", "供应链", "物流", "transportation", "shipping", "warehouse", "仓储"],
    "consulting": ["consulting", "咨询", "advisory", "agency", "服务", "service", "outsourcing", "外包"],
    "media": ["media", "媒体", "marketing", "广告", "ad", "pr", "content", "内容", "entertainment", "娱乐"],
    "realestate": ["real estate", "property", "房地产", "建筑", "construction", "建材"],
    "energy": ["energy", "能源", "oil", "gas", "电力", "power", "新能源", "renewable"],
    "government": ["government", "政府", "ngo", "非营利", "non-profit", "公共部门"],
}

# 公司规模关键词 -> 模板规模代码
COMPANY_SIZE_MAP: Dict[str, str] = {
    "1-10": "1-10", "1 - 10": "1-10", "少于10": "1-10", "<10": "1-10",
    "11-50": "11-50", "11 - 50": "11-50",
    "51-200": "51-200", "51 - 200": "51-200",
    "201-500": "201-500", "201 - 500": "201-500",
    "501-1000": "501-1000", "501 - 1000": "501-1000",
    "1001-5000": "1001-5000", "1001 - 5000": "1001-5000",
    "5001-10000": "5001-10000", "5001 - 10000": "5001-10000",
    "10001+": "10001+", "10000+": "10001+", "10001 +": "10001+",
}


# ==================== 模板加载 ====================

@dataclass
class TemplateRegistry:
    company_headers: List[str] = field(default_factory=list)
    company_industries: Dict[str, str] = field(default_factory=dict)
    company_sizes: Dict[str, str] = field(default_factory=dict)
    task_headers: List[str] = field(default_factory=list)
    skills: List[Dict[str, Any]] = field(default_factory=list)
    skill_aliases: Dict[str, Dict[str, Any]] = field(default_factory=dict)

    def load(self) -> "TemplateRegistry":
        self._load_company_template()
        self._load_skills_template()
        self._load_task_template()
        return self

    def _load_company_template(self):
        if not COMPANY_TEMPLATE_FILE.exists():
            logger.warning(f"公司模板不存在: {COMPANY_TEMPLATE_FILE}")
            return
        wb = load_workbook(COMPANY_TEMPLATE_FILE, read_only=True, data_only=True)
        try:
            ws = wb["平台公司导入"]
            self.company_headers = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            ws_dict = wb["字段字典"]
            for row in ws_dict.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[0]:
                    continue
                field_type, code, display = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip()
                if field_type == "行业":
                    self.company_industries[code] = display
                elif field_type == "公司规模":
                    self.company_sizes[code] = display
        finally:
            wb.close()

    def _load_skills_template(self):
        if not SKILLS_TEMPLATE_FILE.exists():
            logger.warning(f"技能模板不存在: {SKILLS_TEMPLATE_FILE}")
            return
        wb = load_workbook(SKILLS_TEMPLATE_FILE, read_only=True, data_only=True)
        try:
            # 优先读取带三级标签的 "优化-副本" sheet
            ws = wb["优化-副本"] if "优化-副本" in wb.sheetnames else wb["优化"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 6:
                    continue
                # 优化-副本列: 0=排序,1=一级EN,2=一级CN,3=二级EN,4=二级CN,5=三级EN,6=三级CN,7=方向说明
                primary_en = str(row[1]).strip() if row[1] else ""
                primary_cn = str(row[2]).strip() if row[2] else ""
                secondary_en = str(row[3]).strip() if row[3] else ""
                secondary_cn = str(row[4]).strip() if row[4] else ""
                tertiary_en = str(row[5]).strip() if len(row) > 5 and row[5] else secondary_en
                tertiary_cn = str(row[6]).strip() if len(row) > 6 and row[6] else secondary_cn
                note = str(row[7]).strip() if len(row) > 7 and row[7] else ""

                if not primary_en and not primary_cn:
                    continue

                skill = {
                    "primary_en": primary_en,
                    "primary_cn": primary_cn,
                    "secondary_en": secondary_en,
                    "secondary_cn": secondary_cn,
                    "tertiary_en": tertiary_en,
                    "tertiary_cn": tertiary_cn,
                    "note": note,
                }
                self.skills.append(skill)

                # 建立别名索引：同时用中英文各级标签作为匹配关键词
                aliases = [primary_en, primary_cn, secondary_en, secondary_cn, tertiary_en, tertiary_cn, note]
                for alias in aliases:
                    if not alias or alias.lower() in ("none", "nan"):
                        continue
                    key = alias.lower()
                    if key not in self.skill_aliases:
                        self.skill_aliases[key] = skill
        finally:
            wb.close()

    def _load_task_template(self):
        if not TASK_TEMPLATE_FILE.exists():
            logger.warning(f"任务模板不存在: {TASK_TEMPLATE_FILE}")
            return
        wb = load_workbook(TASK_TEMPLATE_FILE, read_only=True, data_only=True)
        try:
            ws = wb.active
            self.task_headers = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        finally:
            wb.close()


# 全局单例
_REGISTRY: Optional[TemplateRegistry] = None


def get_registry() -> TemplateRegistry:
    global _REGISTRY
    if _REGISTRY is None:
        _REGISTRY = TemplateRegistry().load()
    return _REGISTRY


# ==================== 清洗逻辑（复用 reference_spider） ====================

_AD_PATTERNS = [
    re.compile(r"订阅相似职位.*?(?:$|更多职位)", re.I | re.S),
    re.compile(r"更多职位[\s\S]*?(?:$|立即发布)", re.I | re.S),
    re.compile(r"在招人？[\s\S]*?(?:$|关于)", re.I | re.S),
    re.compile(r"解锁有关.*?招聘洞察", re.I | re.S),
    re.compile(r"Chart.*?End of interactive chart\.", re.I | re.S),
    re.compile(r"试用 Premium.*?提醒您。", re.I | re.S),
    re.compile(r"Apply now[\s\S]*?Ready to apply", re.I | re.S),
    re.compile(r"Sign in[\s\S]*?Join now", re.I | re.S),
]


def clean_text(text: Optional[str]) -> str:
    """基础文本清洗：参考 reference_spider 的 _clean_text 逻辑。"""
    if not text:
        return ""
    text = str(text)
    # 解码 HTML 实体
    text = html.unescape(text)
    # 去除 HTML 标签
    text = re.sub(r"<[^>]+>", " ", text)
    # 去除常见广告文案
    for pattern in _AD_PATTERNS:
        text = pattern.sub(" ", text)
    # 统一空白
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_salary_text(salary: Optional[str]) -> str:
    """薪资文本预清洗。"""
    if not salary:
        return ""
    s = clean_text(salary)
    # 移除常见干扰词
    s = re.sub(r"(?i)\b(salary|compensation|pay|range)\b[:\s]*", "", s)
    return s.strip(" :/-–—")


# ==================== 标准化解析函数 ====================

def parse_salary(salary_text: str) -> Dict[str, Any]:
    """解析薪资字符串为结构化字段。

    返回:
        {
            "budget_mode": "custom" | "negotiable",
            "pricing_type": "year" | "month" | "hour" | "day" | "week" | "project" | None,
            "budget_min": float | None,
            "budget_max": float | None,
            "currency": str | None,
            "salary_raw": str,
        }
    """
    result = {
        "budget_mode": "negotiable",
        "pricing_type": None,
        "budget_min": None,
        "budget_max": None,
        "currency": None,
        "salary_raw": salary_text,
    }
    if not salary_text:
        return result

    s = clean_salary_text(salary_text)
    lower = s.lower()

    # 识别常见"面议/negotiable"表述
    if any(k in lower for k in ["negotiable", "competitive", "面议", "另议", "depending", "depends", "flexible"]):
        result["budget_mode"] = "negotiable"
        return result

    # 提取货币
    for key, code in CURRENCY_MAP.items():
        if key in lower or key.upper() in s:
            result["currency"] = code
            # 移除货币符号/文字，方便数字提取
            s = re.sub(re.escape(key), "", s, flags=re.I)
            break

    # 提取周期
    period = None
    period_patterns = [
        (r"(?:per\s*|/|\b)(year|yr|annum|annual)\b", "year"),
        (r"(?:per\s*|/|\b)(month|mo)\b", "month"),
        (r"(?:per\s*|/|\b)(hour|hr)\b", "hour"),
        (r"(?:per\s*|/|\b)(day|daily)\b", "day"),
        (r"(?:per\s*|/|\b)(week|wk)\b", "week"),
        (r"(?:项目总价|固定价格|fixed|project|lump\s*sum)", "project"),
    ]
    for pattern, ptype in period_patterns:
        if re.search(pattern, lower):
            period = ptype
            break

    # 如果包含中文"年薪/月薪/时薪/天/周"
    if not period:
        if "年薪" in salary_text:
            period = "year"
        elif "月薪" in salary_text:
            period = "month"
        elif "时薪" in salary_text:
            period = "hour"
        elif "日薪" in salary_text:
            period = "day"
        elif "周薪" in salary_text:
            period = "week"

    result["pricing_type"] = period
    result["budget_mode"] = "custom" if period else "negotiable"

    # 提取数字区间：支持 10K-20K, $50,000 - $80,000, 15k-25k, 3万-5万 等
    numbers: List[float] = []

    # 匹配 "10K-20K" / "10k-20k" / "$50K-$80K"
    k_matches = re.findall(r"([\d,.]+)\s*[kK]", s)
    for m in k_matches:
        try:
            numbers.append(float(m.replace(",", "")) * 1000)
        except ValueError:
            continue

    # 匹配万（人民币常见）
    wan_matches = re.findall(r"([\d,.]+)\s*[万萬]", s)
    for m in wan_matches:
        try:
            numbers.append(float(m.replace(",", "")) * 10000)
        except ValueError:
            continue

    # 匹配普通数字（带千分位）
    if not numbers:
        plain_matches = re.findall(r"[\d,]+(?:\.\d+)?", s)
        for m in plain_matches:
            try:
                val = float(m.replace(",", ""))
                if val > 100:  # 忽略过小数字（如工作年限）
                    numbers.append(val)
            except ValueError:
                continue

    if numbers:
        result["budget_min"] = min(numbers)
        result["budget_max"] = max(numbers)
        result["budget_mode"] = "custom"

    return result


def parse_country(location_text: Optional[str]) -> Tuple[str, str]:
    """从地点文本解析国家代码和城市。

    返回: (country_code, city)
    """
    if not location_text:
        return ("", "")
    loc = clean_text(location_text)
    lower = loc.lower()

    # 远程优先
    if any(k in lower for k in ["remote", "线上", "远程", "work from home", "wfh"]):
        return ("REMOTE", "")

    # 先匹配国家
    country_code = ""
    for key, code in COUNTRY_MAP.items():
        if re.search(r"\b" + re.escape(key.lower()) + r"\b", lower):
            country_code = code
            # 从文本中移除国家词，剩余部分作为城市
            loc = re.sub(r"\b" + re.escape(key) + r"\b", "", loc, flags=re.I)
            break

    # 提取城市：优先按逗号/竖线/斜杠分隔
    city = ""
    parts = re.split(r"[,|/\\]", loc)
    parts = [p.strip() for p in parts if p.strip()]
    if parts:
        # 取最长或最像城市的片段
        city = max(parts, key=len)

    return (country_code, city)


def parse_work_location_type(text: str) -> str:
    """判断工作方式：online_remote / offline_office。"""
    if not text:
        return ""
    lower = text.lower()
    remote_keywords = ["remote", "online", "线上", "远程", "wfh", "work from home", "anywhere", "home-based"]
    if any(k in lower for k in remote_keywords):
        return "online_remote"
    office_keywords = ["office", "onsite", "on-site", "线下", "现场", "坐班", "hybrid", "混合办公"]
    if any(k in lower for k in office_keywords):
        return "offline_office"
    return "online_remote"  # 海外岗位默认远程可接受


def parse_languages(text: Optional[str]) -> str:
    """从文本提取语言要求，返回逗号分隔的 code。"""
    if not text:
        return ""
    found: List[str] = []
    lower = text.lower()
    for alias, code in LANGUAGE_MAP.items():
        if alias.lower() in lower and code not in found:
            found.append(code)
    return ",".join(found) if found else ""


def parse_posted_date(posted_text: Optional[str]) -> Optional[str]:
    """解析发布时间为 YYYY-MM-DD 或相对日期。"""
    if not posted_text:
        return None
    s = clean_text(posted_text)

    # 绝对日期：Jul 15, 2026 / 2026-07-15 / 2026/07/15 / 2026.07.15
    patterns = [
        r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})",
        r"([A-Za-z]{3,9})\s+(\d{1,2}),\s+(\d{4})",
        r"(\d{1,2})\s+([A-Za-z]{3,9}),\s+(\d{4})",
    ]
    for pattern in patterns:
        m = re.search(pattern, s)
        if m:
            try:
                if pattern.startswith(r"(\d{4})"):
                    year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
                    return date(year, month, day).isoformat()
                else:
                    # 简化为 today if parse fail
                    pass
            except Exception:
                pass

    # 相对日期：x days ago / x天前
    m = re.search(r"(\d+)\s*(天前|days?\s*ago|hours?\s*ago|小时前)", s, re.I)
    if m:
        num = int(m.group(1))
        if "小时" in s or "hour" in s.lower():
            delta = timedelta(hours=num)
        else:
            delta = timedelta(days=num)
        return (datetime.utcnow() - delta).date().isoformat()

    # 中文月日（当年）
    m = re.search(r"(\d{1,2})月(\d{1,2})日", s)
    if m:
        try:
            return date(datetime.utcnow().year, int(m.group(1)), int(m.group(2))).isoformat()
        except Exception:
            pass

    return None


def detect_industry(company_name: str, description: Optional[str]) -> str:
    """根据公司名称和描述匹配行业代码。"""
    text = f"{company_name or ''} {description or ''}".lower()
    scores: Dict[str, int] = {}
    for code, keywords in INDUSTRY_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in text:
                scores[code] = scores.get(code, 0) + len(kw)
    if not scores:
        return "technology"  # 默认值
    return max(scores.items(), key=lambda x: x[1])[0]


def detect_company_size(description: Optional[str]) -> str:
    """从描述中尝试提取公司规模。"""
    if not description:
        return ""
    text = clean_text(description)
    for code in COMPANY_SIZE_MAP:
        if code in text or code.replace("-", " - ") in text:
            return code
    # 尝试匹配 "1000+ employees"
    m = re.search(r"(\d+)\+?\s*(employees|staff|人|员工)", text, re.I)
    if m:
        n = int(m.group(1))
        if n < 11:
            return "1-10"
        elif n < 51:
            return "11-50"
        elif n < 201:
            return "51-200"
        elif n < 501:
            return "201-500"
        elif n < 1001:
            return "501-1000"
        elif n < 5001:
            return "1001-5000"
        elif n < 10001:
            return "5001-10000"
        else:
            return "10001+"
    return ""


def match_skills(text: Optional[str], registry: TemplateRegistry) -> List[Dict[str, Any]]:
    """基于技能库匹配文本中出现的技能标签。"""
    if not text:
        return []
    lower = clean_text(text).lower()
    matched: Dict[str, Dict[str, Any]] = {}
    for alias, skill in registry.skill_aliases.items():
        if alias in lower:
            key = f"{skill['primary_en']}|{skill['secondary_en']}|{skill['tertiary_en']}"
            if key not in matched:
                matched[key] = skill
    return list(matched.values())


# ==================== 单条岗位标准化 ====================

def standardize_job(job: JobRecord, registry: Optional[TemplateRegistry] = None) -> Dict[str, Any]:
    """将 JobRecord 标准化为对齐模板的数据结构。"""
    registry = registry or get_registry()
    raw = job.raw_data or {}

    title = clean_text(job.title or raw.get("title", ""))[:MAX_TITLE_LENGTH]
    company = clean_text(job.company or raw.get("company", ""))[:MAX_COMPANY_LENGTH]
    description = clean_text(job.description or raw.get("description", ""))[:MAX_DESC_LENGTH]
    location_text = clean_text(job.location or raw.get("location", ""))
    salary_text = clean_text(job.salary or raw.get("salary", ""))
    url = job.url or raw.get("url", "")

    # 合并用于技能匹配的文本
    full_text = f"{title} {description} {raw.get('job_category', '')} {raw.get('job_style', '')}"
    skills = match_skills(full_text, registry)

    # 平台公司数据
    country_code, city = parse_country(location_text)
    if country_code == "REMOTE":
        work_location_type = "online_remote"
        country_code = ""
    else:
        work_location_type = parse_work_location_type(f"{location_text} {raw.get('job_style', '')} {description}")

    industry = detect_industry(company, description)
    company_size = detect_company_size(description)

    company_data = {
        "公司中文名（必填，≤200字）": company,
        "公司英文名（可选，≤200字）": company if not _is_chinese_only(company) else "",
        "行业（可选，填写“字段字典”中的行业代码）": industry if industry in registry.company_industries else "technology",
        "国家（可选，默认CN，≤50字）": country_code or "CN",
        "城市（可选，≤50字）": city,
        "公司规模（可选，填写“字段字典”中的规模代码）": company_size,
        "官网（可选，≤255字）": raw.get("company_url", "")[:255],
        "Logo URL（可选，公开http/https图片，jpeg/png/gif/webp，≤1MB）": "",
        "公司简介（可选，≤5000字）": description[:5000],
    }

    # 任务数据
    salary_info = parse_salary(salary_text)
    languages = parse_languages(description)
    posted_date = parse_posted_date(job.raw_data.get("posted_time", "")) if job.raw_data else None
    duration_start = posted_date or datetime.utcnow().date().isoformat()
    duration_end = (datetime.utcnow() + timedelta(days=30)).date().isoformat()

    skill_category = ""
    if skills:
        # 优先使用一级中文分类
        skill_category = skills[0].get("primary_cn") or skills[0].get("primary_en", "")
    if not skill_category:
        skill_category = "销售与商务"

    # 是否长期：没有结束日期或明确长期
    is_long_term = "N"
    if "长期" in description or "long term" in description.lower() or "长期合作" in description:
        is_long_term = "Y"

    task_data = {
        "title（必填，任务标题）": title,
        "external_url": url[:1000],
        "description（必填，任务描述）": description,
        "skill_category（必填：库中分类 ID，或该分类在后台维护的中文名/英文名；示例行取自当前数据库）": skill_category,
        "work_location_type（可选，工作地点类型：online_remote/线上远程 或 offline_office/线下办公）": work_location_type,
        "country（线下办公时必填，国家/地区：代码CN/SG 或 中文名中国/新加坡 或 英文名China/Singapore）": country_code if work_location_type == "offline_office" else "",
        "languages（可选，逗号分隔：languages 表中的 code（如 zh-Hans、en），或表中/常用口语名）": languages,
        "budget_mode（可选，报酬模式：custom/自定义 或 negotiable/另议，默认negotiable）": salary_info["budget_mode"],
        "pricing_type（H=custom时必填：minute/word/hour/day/week/month/year/project）": salary_info["pricing_type"] or "",
        "budget_min（H=custom时可填，最低预算，默认0）": salary_info["budget_min"],
        "budget_max（H=custom时可填，最高预算，须≥budget_min）": salary_info["budget_max"],
        "currency（H=custom时可填，货币：CNY/USD/EUR/GBP/JPY/HKD/SGD/AUD/CAD，...）": salary_info["currency"] or "",
        "duration_start（可选，任务开始日期，格式YYYY-MM-DD，如2026-07-06）": duration_start,
        "duration_end（可选，任务结束日期，格式YYYY-MM-DD，须≥开始日期，如2026-08-05）": duration_end,
        "is_long_term（可选，是否长期任务：Y/是/1=是，N/否/0=否，默认N；填Y则忽略开始结束日期）": is_long_term,
    }

    return {
        "company": company_data,
        "task": task_data,
        "skills": skills,
        "meta": {
            "url_hash": job.url_hash,
            "standardized_at": datetime.utcnow().isoformat(),
        },
    }


def _is_chinese_only(text: str) -> bool:
    """简单判断字符串是否只包含中文字符（用于决定公司名是否放中文列）。"""
    if not text:
        return False
    for ch in text:
        if "\u4e00" <= ch <= "\u9fff":
            continue
        if ch.isalpha() and not ("\u4e00" <= ch <= "\u9fff"):
            return False
    return True


# ==================== 批量标准化服务 ====================

class StandardizerService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.registry = get_registry()

    async def _log(self, task_id: int, level: str, message: str):
        """将清洗/标准化日志写入 CrawlLog。"""
        logger.info(f"[Task {task_id}] {message}")
        self.db.add(CrawlLog(task_id=task_id, level=level, message=message))
        await self.db.commit()

    async def standardize_task_jobs(self, task_id: int) -> Dict[str, int]:
        """对指定任务下所有未标准化的岗位进行标准化。"""
        result = await self.db.execute(
            select(JobRecord).where(
                JobRecord.task_id == task_id,
                JobRecord.is_standardized == False,
            )
        )
        jobs = result.scalars().all()
        processed = updated = 0
        for job in jobs:
            try:
                data = standardize_job(job, self.registry)
                job.standardized_data = data
                job.matched_skills = data["skills"]
                job.is_standardized = True
                job.standardized_at = datetime.utcnow()
                updated += 1
            except Exception as e:
                logger.warning(f"标准化失败 job_id={job.id}: {e}")
            processed += 1
        await self.db.commit()
        await self._log(task_id, "info", f"数据标准化完成: 处理 {processed} 条, 更新 {updated} 条")
        return {"processed": processed, "updated": updated}

    async def standardize_job_by_id(self, job_id: int) -> Optional[JobRecord]:
        job = await self.db.get(JobRecord, job_id)
        if not job:
            return None
        data = standardize_job(job, self.registry)
        job.standardized_data = data
        job.matched_skills = data["skills"]
        job.is_standardized = True
        job.standardized_at = datetime.utcnow()
        await self.db.commit()
        return job

    async def restandardize_all(self, task_id: Optional[int] = None) -> Dict[str, int]:
        """重新标准化（用于规则更新后全量刷新）。"""
        stmt = select(JobRecord)
        if task_id:
            stmt = stmt.where(JobRecord.task_id == task_id)
        result = await self.db.execute(stmt)
        jobs = result.scalars().all()
        updated = 0
        for job in jobs:
            try:
                data = standardize_job(job, self.registry)
                job.standardized_data = data
                job.matched_skills = data["skills"]
                job.is_standardized = True
                job.standardized_at = datetime.utcnow()
                updated += 1
            except Exception as e:
                logger.warning(f"重新标准化失败 job_id={job.id}: {e}")
        await self.db.commit()
        if task_id:
            await self._log(task_id, "info", f"全量重新标准化完成: 共 {len(jobs)} 条, 更新 {updated} 条")
        return {"updated": updated, "total": len(jobs)}


async def run_standardization(db: AsyncSession, task_id: Optional[int] = None) -> Dict[str, int]:
    """对外暴露的标准化入口。"""
    service = StandardizerService(db)
    if task_id:
        return await service.standardize_task_jobs(task_id)
    return await service.restandardize_all()
