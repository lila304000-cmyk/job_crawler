"""Excel 导出服务：严格按照导入模板字段输出。"""

from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models import JobRecord
from services.standardizer import get_registry


async def export_jobs_excel(
    db: AsyncSession,
    task_id: Optional[int] = None,
    standardized_only: bool = True,
) -> bytes:
    """导出岗位为严格匹配导入模板的 Excel 文件。

    Sheet1: 平台公司导入
    Sheet2: 任务导入
    """
    registry = get_registry()

    stmt = select(JobRecord)
    if task_id:
        stmt = stmt.where(JobRecord.task_id == task_id)
    if standardized_only:
        stmt = stmt.where(JobRecord.is_standardized == True)

    result = await db.execute(stmt)
    jobs = result.scalars().all()

    wb = Workbook()

    # ==================== Sheet 1: 平台公司导入 ====================
    ws_company = wb.active
    ws_company.title = "平台公司导入"
    company_headers = registry.company_headers or [
        "公司中文名（必填，≤200字）",
        "公司英文名（可选，≤200字）",
        "行业（可选，填写“字段字典”中的行业代码）",
        "国家（可选，默认CN，≤50字）",
        "城市（可选，≤50字）",
        "公司规模（可选，填写“字段字典”中的规模代码）",
        "官网（可选，≤255字）",
        "Logo URL（可选，公开http/https图片，jpeg/png/gif/webp，≤1MB）",
        "公司简介（可选，≤5000字）",
    ]
    ws_company.append(company_headers)

    seen_companies: Dict[str, Dict[str, Any]] = {}
    for job in jobs:
        data = job.standardized_data or {}
        company = data.get("company", {})
        company_name = company.get("公司中文名（必填，≤200字）", "")
        if not company_name:
            continue
        if company_name not in seen_companies:
            seen_companies[company_name] = company

    for company in seen_companies.values():
        row = [company.get(h, "") for h in company_headers]
        ws_company.append(row)

    # ==================== Sheet 2: 任务导入 ====================
    ws_task = wb.create_sheet(title="任务导入")
    task_headers = registry.task_headers or [
        "title（必填，任务标题）",
        "external_url",
        "description（必填，任务描述）",
        "skill_category（必填：库中分类 ID，或该分类在后台维护的中文名/英文名；示例行取自当前数据库）",
        "work_location_type（可选，工作地点类型：online_remote/线上远程 或 offline_office/线下办公）",
        "country（线下办公时必填，国家/地区：代码CN/SG 或 中文名中国/新加坡 或 英文名China/Singapore）",
        "languages（可选，逗号分隔：languages 表中的 code（如 zh-Hans、en），或表中/常用口语名）",
        "budget_mode（可选，报酬模式：custom/自定义 或 negotiable/另议，默认negotiable）",
        "pricing_type（H=custom时必填：minute/word/hour/day/week/month/year/project）",
        "budget_min（H=custom时可填，最低预算，默认0）",
        "budget_max（H=custom时可填，最高预算，须≥budget_min）",
        "currency（H=custom时可填，货币：CNY/USD/EUR/GBP/JPY/HKD/SGD/AUD/CAD，...）",
        "duration_start（可选，任务开始日期，格式YYYY-MM-DD，如2026-07-06）",
        "duration_end（可选，任务结束日期，格式YYYY-MM-DD，须≥开始日期，如2026-08-05）",
        "is_long_term（可选，是否长期任务：Y/是/1=是，N/否/0=否，默认N；填Y则忽略开始结束日期）",
    ]
    ws_task.append(task_headers)

    for job in jobs:
        data = job.standardized_data or {}
        task = data.get("task", {})
        row = [task.get(h, "") for h in task_headers]
        ws_task.append(row)

    # ==================== 样式 ====================
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ws in [ws_company, ws_task]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        # 自动列宽（简单实现）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
